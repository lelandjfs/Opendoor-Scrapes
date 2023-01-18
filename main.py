import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook


headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'}


dealDataList = []

def getData(location, page):
    
    bread_crumb_api = f"https://buy.opendoor.com/zones/from_breadcrumb.json?breadcrumb_path=%2F{location}"
    r = requests.get(bread_crumb_api, headers = headers)
    json_data = r.json()
    bounds_json = json_data['bounds']
    centers = json_data['center']
    
    bound_north = bounds_json[0][0]
    bound_west = bounds_json[0][1]
    bound_south = bounds_json[1][0]
    bound_east = bounds_json[1][1]
    
    listing_api = f"https://buy.opendoor.com/zones/null/list_properties.json?page={page}&page_size=30&sort=newest&properties_filter%5Binclude_homebuilder%5D=true&include_markers=1000&bounds%5Bnorth%5D={bound_north}&bounds%5Beast%5D={bound_east}&bounds%5Bsouth%5D={bound_south}&bounds%5Bwest%5D={bound_west}&location%5B%5D={centers[1]}&location%5B%5D={centers[0]}"
    
    response = requests.get(listing_api)
    
    page_data_list = []
        
    for item in response.json()['properties']:
        deal = {
            'location' : location,
            'address': item['building_address'],
            'price' :  item['current_list_price'],
            'bathrooms': item['bathrooms'],
            'bedrooms': item['bedrooms'],
            'sqft': item['sqft'],
            'realtor': item['listing_office'],
            'date scraped' : datetime.now().strftime("%Y-%m-%d")
        }
        page_data_list.append(deal)
        
    return page_data_list

# will need to have function showing how many pages for the 2nd range number below in future

for x in range(1,100):
    dealDataList = dealDataList + getData('sacramento', x)


df = pd.DataFrame(dealDataList)
current_date = datetime.now().strftime("%Y-%m-%d")
df.to_excel(r'C:\Users\lelan\OneDrive\Documents\Python\Web Data Scraper\OpenDoor\Sacramento\Archive\opendoorresults_' + current_date + '.xlsx', index = False)
# if doesnt work above, use 2 \\ for every 1 currently
print('Done.')

# BREAK - load file into master

source_file = load_workbook(r'C:\Users\lelan\OneDrive\Documents\Python\Web Data Scraper\OpenDoor\Sacramento\Archive\opendoorresults_' + current_date + '.xlsx')
dest_file = load_workbook(r'C:\Users\lelan\OneDrive\Documents\Python\Web Data Scraper\OpenDoor\Sacramento\Master\OpenDoorMaster.xlsx')

source_sheet = source_file.active
dest_sheet = dest_file.active

max_row = source_sheet.max_row
last_row = dest_sheet.max_row

for i in range(2, max_row + 1): # start loop from second row
    for j in range(1, source_sheet.max_column + 1):
        dest_sheet.cell(row=last_row + i - 1, column=j).value = source_sheet.cell(row=i, column=j).value


dest_file.save(r'C:\Users\lelan\OneDrive\Documents\Python\Web Data Scraper\OpenDoor\Sacramento\Master\OpenDoorMaster.xlsx')
dest_file.close()
source_file.close()

print ('Saved to master')
